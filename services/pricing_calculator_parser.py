#!/usr/bin/env python3
"""
Huawei Cloud Pricing Calculator Quotation Parser (C&C v3.0 format)
Handles the multi-sheet Excel export from https://www.huaweicloud.com/intl/en-us/pricing/calculator.html

This is a NEW standalone parser — does NOT modify or depend on existing parsers.
Format detection is done by checking for characteristic sheet names and column layouts.
"""

import openpyxl
import re
import json
from typing import Optional, Dict, Any, List, Tuple


# ── Format Detection ──────────────────────────────────────────────

def is_pricing_calculator_format(file_path: str) -> bool:
    """
    Detect if the file is a Huawei Cloud Pricing Calculator export.
    Checks for characteristic sheet names and the 12-column layout.
    """
    if not str(file_path).lower().endswith(('.xlsx', '.xls')):
        return False
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        
        # Must have "Price Calculator Summary" or similar summary sheet
        has_summary = any('price calculator summary' in s for s in sheet_names)
        if not has_summary:
            # Check for 'Global Summary' text in any sheet
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
                    for cell in row:
                        if cell and 'global summary' in str(cell).lower():
                            wb.close()
                            return True
        
        # Must have service sheets with characteristic names
        service_sheet_keywords = ['compute - prod', 'compute - dev', 'storage', 'connectivity', 'security', 'support']
        service_sheets = 0
        for keyword in service_sheet_keywords:
            if any(keyword in s for s in sheet_names):
                service_sheets += 1
        
        wb.close()
        
        # If we have summary + at least 3 service sheets, it's a pricing calculator export
        if has_summary and service_sheets >= 3:
            return True
        if service_sheets >= 4:
            return True
            
        return False
        
    except Exception:
        return False


def is_pricing_calculator_url(url: str) -> bool:
    """Detect if a URL is a Huawei Cloud Pricing Calculator share link."""
    if not url:
        return False
    url_lower = url.lower()
    return ('huaweicloud.com' in url_lower and 
            ('pricing/calculator' in url_lower or 'calculator.html' in url_lower) and
            ('sharelistid=' in url_lower or 'shareListId=' in url_lower))


# ── Sheet-Level Parsing ───────────────────────────────────────────

# Standard columns in Pricing Calculator export (header at row 8, data from row 9)
PC_COLUMNS = [
    'required',       # 0: Customer's server name/role description
    'service',        # 1: Formal Huawei service name (e.g., "Data Warehouse Service 1")
    'description',    # 2: Human-readable description
    'region',         # 3: Huawei Cloud region
    'az',             # 4: Availability Zone
    'billing_mode',   # 5: Monthly / RI / Pay-per-use
    'purchase_amount',# 6: Numeric purchase amount
    'unit',           # 7: Unit (month, hours, days, etc.)
    'quantity',       # 8: Quantity
    'specifications', # 9: Detailed specs string
    'unit_price',     # 10: Unit Price in USD
    'monthly_price'   # 11: Monthly Price in USD
]


def _parse_pricing_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> Dict[str, Any]:
    """
    Parse a single service sheet from the Pricing Calculator export.
    Returns a dict with items list, total price, and metadata.
    """
    items = []
    total_price = 0.0
    
    # Find the header row (containing 'Service' and 'Description')
    header_row = None
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or '').strip().lower() 
                    for c in range(1, ws.max_column + 1)]
        if 'service' in row_vals and 'description' in row_vals:
            header_row = row_idx
            break
    
    if header_row is None:
        # Fallback: try row 8 (standard position)
        header_row = 8
    
    # Build column index map from header
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
        if 'service' in val and 'description' not in val:
            col_map['service'] = c
        elif 'description' in val:
            col_map['description'] = c
        elif 'billing mode' in val:
            col_map['billing_mode'] = c
        elif 'purchase amount' in val:
            col_map['purchase_amount'] = c
        elif 'unit' in val and 'unit price' not in val:
            col_map['unit'] = c
        elif 'quantity' in val:
            col_map['quantity'] = c
        elif 'specifications' in val:
            col_map['specifications'] = c
        elif 'unit price' in val:
            col_map['unit_price'] = c
        elif 'monthly price' in val:
            col_map['monthly_price'] = c
        elif 'required' in val:
            col_map['required'] = c
        elif 'region' in val:
            col_map['region'] = c
        elif 'az' in val:
            col_map['az'] = c
    
    # Read data rows (after header)
    data_start = header_row + 1
    for row_idx in range(data_start, ws.max_row + 1):
        service = str(ws.cell(row=row_idx, column=col_map.get('service', 2)).value or '').strip()
        description = str(ws.cell(row=row_idx, column=col_map.get('description', 3)).value or '').strip()
        required = str(ws.cell(row=row_idx, column=col_map.get('required', 1)).value or '').strip()
        
        # Skip empty rows (no service AND no required name)
        if not service and not required:
            continue
        if service.lower() in ['nan', 'none', ''] and required.lower() in ['nan', 'none', '']:
            continue
        
        # Skip summary/footer rows — check BOTH service and required columns
        skip_lower = (service + ' ' + required).lower()
        if 'total price' in skip_lower:
            try:
                total_price = float(str(ws.cell(row=row_idx, column=col_map.get('monthly_price', 12)).value or 0))
            except (ValueError, TypeError):
                pass
            continue
        if 'preceding price' in skip_lower or 'monthly price' in skip_lower:
            continue
        
        # Skip rows where service is just a number (summary row like "1094.26" in service col)
        if service.replace('.', '').replace(',', '').isdigit():
            continue
        
        # Sub-item detection: if required is empty but service has data, it's a child item
        # (e.g., EVS disk attached to a server). Attach to previous item.
        is_sub_item = bool(not required and service)
        billing_mode = str(ws.cell(row=row_idx, column=col_map.get('billing_mode', 6)).value or '').strip()
        purchase_amount = str(ws.cell(row=row_idx, column=col_map.get('purchase_amount', 7)).value or '').strip()
        unit = str(ws.cell(row=row_idx, column=col_map.get('unit', 8)).value or '').strip()
        quantity = str(ws.cell(row=row_idx, column=col_map.get('quantity', 9)).value or '').strip()
        specs = str(ws.cell(row=row_idx, column=col_map.get('specifications', 10)).value or '').strip()
        region = str(ws.cell(row=row_idx, column=col_map.get('region', 4)).value or '').strip()
        az = str(ws.cell(row=row_idx, column=col_map.get('az', 5)).value or '').strip()
        
        try:
            unit_price = float(str(ws.cell(row=row_idx, column=col_map.get('unit_price', 11)).value or 0))
        except (ValueError, TypeError):
            unit_price = 0.0
        
        try:
            monthly_price = float(str(ws.cell(row=row_idx, column=col_map.get('monthly_price', 12)).value or 0))
        except (ValueError, TypeError):
            monthly_price = 0.0
        
        try:
            qty = int(float(quantity)) if quantity else 1
        except (ValueError, TypeError):
            qty = 1
        
        item_data = {
            'name': required,         # Human-readable name (was 'Required' column in PC)
            'type': service,          # Formal service name (e.g., "Elastic Cloud Server 1")
            'description': description,
            'region': region,
            'az': az,
            'billing_mode': billing_mode,
            'purchase_amount': purchase_amount,
            'unit': unit,
            'quantity': qty,
            'specifications': specs,
            'unit_price': unit_price,
            'monthly_price': monthly_price,
            'sub_items': []
        }
        
        if is_sub_item:
            # Attach to previous item as a sub-item (e.g., EVS disk on a server)
            if items:
                items[-1]['sub_items'].append(item_data)
                items[-1]['monthly_price'] += monthly_price
            else:
                items.append(item_data)
        else:
            items.append(item_data)
    
    # If we didn't find a total row, sum up the monthly prices
    if total_price == 0.0 and items:
        total_price = sum(item['monthly_price'] for item in items)
    
    return {
        'items': items,
        'total_price': total_price,
        'item_count': len(items)
    }


# ── Category Classification ───────────────────────────────────────

def _classify_item(item: Dict[str, Any]) -> str:
    """
    Classify a parsed line item into its resource category.
    Returns one of: compute, database, storage, networking, security, support, unknown
    """
    svc_type = item.get('type', '').lower()
    desc = item.get('description', '').lower()
    name = item.get('name', '').lower()
    combined = f"{svc_type} {desc} {name}"
    
    # Compute services
    if any(x in combined for x in ['elastic cloud server', 'ecs', 'bare metal', 'bms', 
                                     'flexus x instance', 'elastic load balance', 'elb']):
        return 'compute'
    
    # Database / Data Warehouse
    if any(x in combined for x in ['data warehouse service', 'dws', 'relational database',
                                     'rds', 'gaussdb', 'document database', 'dds', 
                                     'redis', 'dcs', 'gauss']):
        return 'database'
    
    # Storage
    if any(x in combined for x in ['elastic volume service', 'evs', 'object storage', 'obs',
                                     'cloud backup', 'cbr', 'sfs', 'scalable file',
                                     'volume']):
        return 'storage'
    
    # Networking
    if any(x in combined for x in ['virtual private network', 'vpn', 'nat gateway',
                                     'elastic ip', 'eip', 'vpc', 'direct connect',
                                     'cdn', 'load balance']):
        if 'elastic load balance' not in combined:
            return 'networking'
        return 'compute'  # ELB goes to compute
    
    # Security
    if any(x in combined for x in ['host security', 'hss', 'web application firewall',
                                     'waf', 'anti-ddos', 'bastion', 'firewall']):
        return 'security'
    
    # Support
    if any(x in combined for x in ['support plan', 'support plans']):
        return 'support'
    
    return 'unknown'


# ── Spec Parsing ──────────────────────────────────────────────────

def _parse_specs_for_topology(item: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract technical specs (vCPUs, RAM, storage, OS) from the specifications string.
    """
    specs = item.get('specifications', '')
    description = item.get('description', '')
    svc_type = item.get('type', '')
    combined = f"{specs} {description}"
    
    result = {
        'vcpus': 0,
        'ram_gb': 0,
        'storage_gb': 0,
        'os': 'Unknown',
        'instance_type': 'Unknown'
    }
    
    # Extract vCPUs
    vcpu_patterns = [
        r'(\d+)\s*vCPU',
        r'x(\d+)\.\d+u',
        r'(\d+)\s*cores?',
        r'(\d+)\s*vCPUs',
    ]
    for pat in vcpu_patterns:
        m = re.search(pat, specs, re.IGNORECASE)
        if m:
            result['vcpus'] = int(m.group(1))
            break
    
    # Extract RAM
    ram_patterns = [
        r'(\d+)\s*GiB',
        r'(\d+)\s*GB',
        r'x\d+\.\d+u\.(\d+)g',
        r'(\d+)\s*GB\s*RAM',
    ]
    for pat in ram_patterns:
        m = re.search(pat, specs, re.IGNORECASE)
        if m:
            result['ram_gb'] = int(m.group(1))
            break
    
    # Extract OS
    os_patterns = [
        r'(Huawei Cloud EulerOS[^;|]*)',
        r'(CentOS[^;|]*)',
        r'(Windows[^;|]*)',
        r'(Ubuntu[^;|]*)',
        r'(Red Hat[^;|]*)',
        r'(Debian[^;|]*)',
        r'(AlmaLinux[^;|]*)',
        r'(Oracle[^;|]*)',
        r'(openSUSE[^;|]*)',
    ]
    for pat in os_patterns:
        m = re.search(pat, specs, re.IGNORECASE)
        if m:
            result['os'] = m.group(1).strip()
            break
    
    # Extract storage
    storage_sum = 0
    for m in re.finditer(r'(?:SSD|SAS|SATA|Disk|GP SSD)[^|]*\|\s*(\d+)\s*GiB', specs, re.IGNORECASE):
        storage_sum += int(m.group(1))
    for m in re.finditer(r'(\d+)\s*GB\s*\*\s*\d+\s*Node', specs, re.IGNORECASE):
        # e.g., "250 GB * 3 Node" -> 750 GB total
        parts = re.match(r'(\d+)\s*GB\s*\*\s*(\d+)\s*Node', m.group(0), re.IGNORECASE)
        if parts:
            storage_sum += int(parts.group(1)) * int(parts.group(2))
    if storage_sum > 0:
        result['storage_gb'] = storage_sum
    else:
        # Fallback: any standalone GB number
        fallback = re.findall(r'(\d+)\s*GB', specs, re.IGNORECASE)
        for fb in fallback:
            fb_val = int(fb)
            if fb_val != result['ram_gb'] and fb_val < 50000:  # filter out extreme values
                result['storage_gb'] = fb_val
                break
    
    # Extract instance type
    type_patterns = [
        r'General computing-plus',
        r'General computing',
        r'Memory-optimized',
        r'Compute-optimized',
        r'Flexus X Instance',
        r'x\d+\.\d+u\.\d+g',
        r'm\d+\.\w+\.\d+',
        r'c\d+\.\w+\.\d+',
        r'd\d+\.\w+\.\d+',
    ]
    for pat in type_patterns:
        m = re.search(pat, specs, re.IGNORECASE)
        if m:
            result['instance_type'] = m.group(0).strip()
            break
    
    return result


# ── Main Parser Entry Point ───────────────────────────────────────

def parse_pricing_calculator(file_path: str, customer_name: str = "TBD_Customer") -> Dict[str, Any]:
    """
    Parse a Huawei Cloud Pricing Calculator export (C&C v3.0 format).
    Returns the standard blueprint.json schema.
    """
    print(f"🔄 Parsing Pricing Calculator export: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Initialize blueprint (standard schema)
    blueprint = {
        "customer": customer_name,
        "delivery_scope": "landing_zone_only",
        "governance": {"requires_hypercare": False, "maintenance_windows": []},
        "topology": {
            "network": [],
            "compute": [],
            "databases": [],
            "storage": [],
            "security": []
        },
        "commercial_intent": {
            "deployable_assets": [],
            "account_assets": [],
            "pricing_summary": {
                "ppu_total": 0.0,
                "ri_total": 0.0,
                "grand_total": 0.0,
                "currency": "USD",
                "category_totals": {}
            }
        },
        "format": "pricing_calculator_v3",
        "sheets_parsed": []
    }
    
    # Identify service sheets (any sheet with the 12-column layout)
    service_sheet_names = []
    for name in wb.sheetnames:
        # Skip metadata sheets
        if name.lower() in ['discount', 'cover', 'disclaimer', 'solution introduction', 'tmp_v']:
            continue
        # Check if it has the characteristic header
        ws = wb[name]
        for row_idx in range(1, min(ws.max_row + 1, 25)):
            row_vals = [str(ws.cell(row=row_idx, column=c).value or '').strip().lower()
                        for c in range(1, min(ws.max_column + 1, 13))]
            if 'service' in row_vals and 'description' in row_vals and any('monthly price' in v for v in row_vals):
                service_sheet_names.append(name)
                break
    
    # Parse each service sheet
    all_items = []
    category_totals = {}
    
    for sheet_name in service_sheet_names:
        ws = wb[sheet_name]
        sheet_data = _parse_pricing_sheet(ws, sheet_name)
        
        if sheet_data['items']:
            blueprint['sheets_parsed'].append({
                'name': sheet_name,
                'items': sheet_data['item_count'],
                'total': sheet_data['total_price']
            })
            
            for item in sheet_data['items']:
                item['source_sheet'] = sheet_name
                cat = _classify_item(item)
                item['category'] = cat
                
                # Accumulate category totals
                if cat not in category_totals:
                    category_totals[cat] = 0.0
                category_totals[cat] += item['monthly_price']
                
                # Track PPU vs RI
                billing = item.get('billing_mode', '').lower()
                if any(x in billing for x in ['ri', 'reserved']):
                    blueprint['commercial_intent']['pricing_summary']['ri_total'] += item['monthly_price']
                else:
                    blueprint['commercial_intent']['pricing_summary']['ppu_total'] += item['monthly_price']
                
                blueprint['commercial_intent']['pricing_summary']['grand_total'] += item['monthly_price']
                
            all_items.extend(sheet_data['items'])
    
    blueprint['commercial_intent']['pricing_summary']['category_totals'] = category_totals
    
    # Read the summary sheet for cross-reference
    for name in wb.sheetnames:
        if 'price calculator summary' in name.lower():
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                vals = [str(c).strip() if c else '' for c in row]
                if 'monthly total' in vals[1].lower() if len(vals) > 1 else '':
                    try:
                        blueprint['commercial_intent']['pricing_summary']['summary_total'] = float(vals[2])
                    except (ValueError, IndexError):
                        pass
                    break
            break
    
    # Build topology from items by category
    for item in all_items:
        cat = item.get('category', 'unknown')
        item_name = item.get('name', item.get('description', 'unnamed'))
        service_type = item.get('type', 'Unknown')
        
        if cat == 'compute':
            parsed = _parse_specs_for_topology(item)
            blueprint["topology"]["compute"].append({
                "name": _clean_name(item_name),
                "flavor": parsed['instance_type'] if parsed['instance_type'] != 'Unknown' else service_type,
                "is_public": False,
                "status": "OK" if parsed['vcpus'] > 0 else "WARNING",
                "metadata": {
                    "tier": item.get('source_sheet', ''),
                    "os_type": parsed['os'],
                    "cpu_cores": parsed['vcpus'],
                    "ram_gb": parsed['ram_gb'],
                    "storage_gb": parsed['storage_gb'],
                    "billing_mode": item.get('billing_mode', ''),
                    "monthly_price": item.get('monthly_price', 0)
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'database':
            # DWS clusters
            parsed = _parse_specs_for_topology(item)
            blueprint["topology"]["databases"].append({
                "name": _clean_name(item_name),
                "engine": service_type,
                "version": "Unknown",
                "status": "OK",
                "metadata": {
                    "specs": item.get('specifications', ''),
                    "vcpus": parsed['vcpus'],
                    "ram_gb": parsed['ram_gb'],
                    "storage_gb": parsed['storage_gb'],
                    "monthly_price": item.get('monthly_price', 0)
                }
            })
        elif cat == 'networking':
            net_type = _classify_network_type(item)
            blueprint["topology"]["network"].append({
                "name": _clean_name(item_name),
                "type": net_type,
                "cidr": "N/A",
                "status": "OK",
                "metadata": {
                    "monthly_price": item.get('monthly_price', 0),
                    "billing_mode": item.get('billing_mode', '')
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'storage':
            st_type = _classify_storage_type(item)
            blueprint["topology"]["storage"].append({
                "name": _clean_name(item_name),
                "type": st_type,
                "location": item.get('region', 'Global'),
                "status": "OK",
                "metadata": {
                    "specs": item.get('specifications', ''),
                    "monthly_price": item.get('monthly_price', 0)
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'security':
            blueprint["topology"]["security"].append({
                "name": _clean_name(item_name),
                "type": service_type,
                "status": "OK",
                "metadata": {
                    "monthly_price": item.get('monthly_price', 0),
                    "billing_mode": item.get('billing_mode', ''),
                    "quantity": item.get('quantity', 1)
                }
            })
        
        # Add to commercial intent assets
        asset = {
            "id": f"{cat}_{len(blueprint['commercial_intent']['deployable_assets'])}",
            "type": service_type,
            "name": item_name,
            "billing_mode": item.get('billing_mode', 'Pay-per-use'),
            "specification": item.get('specifications', '')[:100],
            "unit_price": item.get('unit_price', 0),
            "monthly_price": item.get('monthly_price', 0),
            "category": cat,
            "source_sheet": item.get('source_sheet', '')
        }
        
        if cat in ['compute', 'database', 'networking', 'storage']:
            blueprint['commercial_intent']['deployable_assets'].append(asset)
        else:
            blueprint['commercial_intent']['account_assets'].append(asset)
    
    wb.close()
    
    # Summary stats
    print(f"✅ Pricing Calculator parse complete:")
    print(f"   📊 {len(all_items)} line items across {len(service_sheet_names)} sheets")
    print(f"   💰 Grand Total: ${blueprint['commercial_intent']['pricing_summary']['grand_total']:.2f} USD/mo")
    print(f"   📋 Categories: {json.dumps(category_totals, indent=2)}")
    
    return blueprint


# ── Shared Topology Builder ──────────────────────────────────────

def _build_topology_from_items(all_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Build a topology dict from parsed line items.
    Used by both XLSX and API-based parsers.
    """
    topology = {
        "network": [],
        "compute": [],
        "databases": [],
        "storage": [],
        "security": []
    }
    
    for item in all_items:
        cat = item.get('category', _classify_item(item))
        item_name = item.get('name', item.get('description', 'unnamed'))
        service_type = item.get('type', 'Unknown')
        
        if cat == 'compute':
            parsed = _parse_specs_for_topology(item)
            topology["compute"].append({
                "name": _clean_name(item_name),
                "flavor": parsed['instance_type'] if parsed['instance_type'] != 'Unknown' else service_type,
                "is_public": False,
                "status": "OK" if parsed['vcpus'] > 0 else "WARNING",
                "metadata": {
                    "tier": item.get('source_sheet', ''),
                    "os_type": parsed['os'],
                    "cpu_cores": parsed['vcpus'],
                    "ram_gb": parsed['ram_gb'],
                    "storage_gb": parsed['storage_gb'],
                    "billing_mode": item.get('billing_mode', ''),
                    "monthly_price": item.get('monthly_price', 0)
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'database':
            parsed = _parse_specs_for_topology(item)
            topology["databases"].append({
                "name": _clean_name(item_name),
                "engine": service_type,
                "version": "Unknown",
                "status": "OK",
                "metadata": {
                    "specs": item.get('specifications', ''),
                    "vcpus": parsed['vcpus'],
                    "ram_gb": parsed['ram_gb'],
                    "storage_gb": parsed['storage_gb'],
                    "monthly_price": item.get('monthly_price', 0)
                }
            })
        elif cat == 'networking':
            net_type = _classify_network_type(item)
            topology["network"].append({
                "name": _clean_name(item_name),
                "type": net_type,
                "cidr": "N/A",
                "status": "OK",
                "metadata": {
                    "monthly_price": item.get('monthly_price', 0),
                    "billing_mode": item.get('billing_mode', '')
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'storage':
            st_type = _classify_storage_type(item)
            topology["storage"].append({
                "name": _clean_name(item_name),
                "type": st_type,
                "location": item.get('region', 'Global'),
                "status": "OK",
                "metadata": {
                    "specs": item.get('specifications', ''),
                    "monthly_price": item.get('monthly_price', 0)
                },
                "sub_items": item.get('sub_items', [])
            })
        elif cat == 'security':
            topology["security"].append({
                "name": _clean_name(item_name),
                "type": service_type,
                "status": "OK",
                "metadata": {
                    "monthly_price": item.get('monthly_price', 0),
                    "billing_mode": item.get('billing_mode', ''),
                    "quantity": item.get('quantity', 1)
                }
            })
    
    return topology


# ── URL-based Import ──────────────────────────────────────────────

# API endpoint for Huawei Cloud Pricing Calculator share data
PRICING_API_BASE = "https://portal-intl.huaweicloud.com"
PRICING_SHARE_PATH = "/api/calculator/rest/cbc/portalcalculatornodeservice/v4/api/share/detail"


def _fetch_share_api_data(share_id: str, language: str = "en-us") -> Dict[str, Any]:
    """
    Call the Huawei Cloud Pricing Calculator API to fetch shared cart data.
    
    Args:
        share_id: The shareListId from the calculator URL
        language: Language code (e.g., 'en-us', 'es-mx', 'pt-br')
    
    Returns:
        Parsed JSON response from the API
    
    Raises:
        ValueError: If the API returns an error or the data is invalid
        ConnectionError: If the API cannot be reached
    """
    import urllib.request
    import urllib.error
    
    api_url = f"{PRICING_API_BASE}{PRICING_SHARE_PATH}?key={share_id}&language={language}"
    
    try:
        req = urllib.request.Request(api_url)
        req.add_header('Accept', 'application/json')
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; ERP-Migration-Factory/1.0)')
        
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode('utf-8')
            data = json.loads(raw)
    except urllib.error.HTTPError as e:
        raise ConnectionError(f"Pricing API returned HTTP {e.code}: {e.reason}")
    except urllib.error.URLError as e:
        raise ConnectionError(f"Cannot reach Pricing API: {e.reason}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Pricing API returned invalid JSON: {e}")
    
    if data.get('retCode') != '200':
        raise ValueError(f"Pricing API error: {data.get('retDesc', 'Unknown error')} (code={data.get('retCode')})")
    
    return data


def _parse_cart_to_items(cart_data: List[Dict[str, Any]], currency: str = "USD") -> List[Dict[str, Any]]:
    """
    Convert Huawei Cloud Pricing Calculator cartListData to our standard item format.
    
    Each cart entry becomes one or more line items with fields:
        name, type, specifications, unit_price, monthly_price, quantity, 
        billing_mode, region, source_sheet
    """
    items = []
    
    for cart_entry in cart_data:
        sp = cart_entry.get('selectedProduct', {})
        product_name = sp.get('description', 'Unnamed Product')
        service_code = sp.get('serviceCode', 'unknown')
        region = sp.get('region', '')
        charge_mode = sp.get('chargeMode', 'Pay-per-use')
        charge_mode_name = sp.get('chargeModeName', charge_mode)
        
        # Map charge mode to standardized names
        billing_mode_map = {
            'PERIOD': 'Yearly/Monthly',
            'ON_DEMAND': 'Pay-per-use',
            'MONTHLY': 'Yearly/Monthly',
            'YEARLY': 'Yearly/Monthly',
        }
        billing_mode = billing_mode_map.get(charge_mode, charge_mode_name)
        
        line_items = sp.get('productAllInfos', [])
        
        for info in line_items:
            ir = info.get('inquiryResult', {})
            qty = info.get('productNum', 1)
            unit_price = info.get('amount', 0) or 0
            total_price = ir.get('amount', 0) or 0
            
            # Build specification string
            spec_parts = []
            spec_code = info.get('resourceSpecCode', '')
            spec_desc = info.get('productSpecSysDesc', '')
            if spec_code:
                spec_parts.append(spec_code)
            if spec_desc and spec_desc != spec_code:
                spec_parts.append(spec_desc)
            # Add SKU info
            for sku in info.get('_skuInfo', [])[:3]:
                if sku and sku not in spec_desc:
                    spec_parts.append(sku)
            
            specifications = ' | '.join(spec_parts) if spec_parts else f"{product_name} configuration"
            
            item = {
                'name': product_name,
                'type': service_code,
                'specifications': specifications,
                'unit_price': unit_price,
                'monthly_price': total_price,
                'quantity': qty,
                'billing_mode': billing_mode,
                'region': region,
                'source_sheet': f'API:{service_code}',
            }
            items.append(item)
    
    return items


def parse_from_share_url(url: str, customer_name: str = "TBD_Customer") -> Dict[str, Any]:
    """
    Parse quotation data from a Huawei Cloud Pricing Calculator share URL.
    
    Calls the official Huawei Cloud Pricing Calculator API to retrieve the full
    shared cart data (same data displayed in the browser calculator), then converts
    it to the standard ERP Migration Factory blueprint format — no manual upload needed.
    
    API: GET portal-intl.huaweicloud.com/api/calculator/rest/cbc/portalcalculatornodeservice/v4/api/share/detail?key=<shareListId>&language=<lang>
    """
    if not is_pricing_calculator_url(url):
        raise ValueError(f"URL is not a valid Huawei Cloud Pricing Calculator share link: {url}")
    
    # Extract shareListId from URL
    match = re.search(r'shareListId=([a-f0-9]+)', url, re.IGNORECASE)
    share_id = match.group(1) if match else None
    if not share_id:
        raise ValueError(f"Cannot extract shareListId from URL: {url}")
    
    # Extract currency from URL
    currency_match = re.search(r'currentCurrency=([A-Z]{3})', url, re.IGNORECASE)
    currency = currency_match.group(1) if currency_match else 'USD'
    
    # Extract language from URL path (e.g., /intl/en-us/ → en-us, /cn/ → zh-cn)
    lang_match = re.search(r'/(?:intl/)?([a-z]{2}-[a-z]{2})/', url, re.IGNORECASE)
    language = lang_match.group(1) if lang_match else 'en-us'
    
    if not lang_match:
        # Try single-segment locale (e.g., /es/ → es, /pt/ → pt)
        simple_match = re.search(r'/([a-z]{2})/', url, re.IGNORECASE)
        if simple_match and simple_match.group(1).lower() not in ('cn', 'us', 'sg', 'hk'):
            language = simple_match.group(1).lower() + '-' + simple_match.group(1).lower()
        # else keep en-us default
    
    print(f"📡 Fetching shared cart data from Huawei Cloud API...")
    print(f"   Share ID: {share_id}")
    print(f"   Language: {language}, Currency: {currency}")
    
    # Step 1: Fetch data from the Pricing Calculator API
    try:
        api_data = _fetch_share_api_data(share_id, language)
    except (ConnectionError, ValueError) as e:
        raise ValueError(f"Failed to fetch pricing data from Huawei Cloud: {e}")
    
    cart_data = api_data.get('data', {}).get('cartListData', [])
    if not cart_data:
        raise ValueError("Shared cart is empty or API returned no data")
    
    print(f"   ✅ Received {len(cart_data)} products from shared cart")
    
    # Step 2: Convert cart data to standardized item format
    all_items = _parse_cart_to_items(cart_data, currency)
    print(f"   ✅ Parsed {len(all_items)} line items")
    
    # Step 3: Classify items and build topology + commercial intent
    # Reuse the same classification logic as the XLSX parser
    categories = {
        'compute': [],
        'database': [],
        'storage': [],
        'networking': [],
        'security': [],
        'support': [],
        'other': []
    }
    
    for item in all_items:
        cat = _classify_item(item)
        if cat not in categories:
            cat = 'other'
        categories[cat].append(item)
    
    # Log category distribution
    category_counts = {k: len(v) for k, v in categories.items() if v}
    print(f"   📊 Categories: {json.dumps(category_counts)}")
    
    # Step 4: Build topology
    topology = _build_topology_from_items(all_items)
    
    # Step 5: Build commercial intent assets
    deployable_assets = []
    account_assets = []
    
    for item in all_items:
        cat = _classify_item(item)
        item_name = item.get('name', 'unnamed')
        
        asset = {
            "id": f"{cat}_{len(deployable_assets) if cat in ['compute','database','networking','storage'] else len(account_assets)}",
            "type": item.get('type', 'unknown'),
            "name": item_name,
            "billing_mode": item.get('billing_mode', 'Pay-per-use'),
            "specification": item.get('specifications', '')[:100],
            "unit_price": item.get('unit_price', 0),
            "monthly_price": item.get('monthly_price', 0),
            "category": cat,
            "source_sheet": item.get('source_sheet', 'API'),
        }
        
        if cat in ['compute', 'database', 'networking', 'storage']:
            deployable_assets.append(asset)
        else:
            account_assets.append(asset)
    
    # Step 6: Calculate pricing totals
    ppu_total = sum(
        i.get('monthly_price', 0) for i in all_items 
        if i.get('billing_mode') in ['Pay-per-use', 'Pay per Use', 'ON_DEMAND']
    )
    ri_total = sum(
        i.get('monthly_price', 0) for i in all_items 
        if i.get('billing_mode') in ['Yearly/Monthly', 'Reserved Instance', 'PERIOD']
    )
    grand_total = sum(i.get('monthly_price', 0) for i in all_items)
    
    # Step 7: Assemble blueprint
    blueprint = {
        "customer": customer_name,
        "delivery_scope": "landing_zone_only",
        "governance": {"requires_hypercare": False, "maintenance_windows": []},
        "topology": topology,
        "commercial_intent": {
            "deployable_assets": deployable_assets,
            "account_assets": account_assets,
            "pricing_summary": {
                "ppu_total": round(ppu_total, 2),
                "ri_total": round(ri_total, 2),
                "grand_total": round(grand_total, 2),
                "currency": currency
            }
        },
        "format": "pricing_calculator_api",
        "source_url": url,
        "share_list_id": share_id,
        "api_version": "v4",
        "product_count": len(cart_data),
        "item_count": len(all_items),
    }
    
    print(f"   💰 Grand Total: ${grand_total:.2f} {currency}/mo")
    print(f"   📋 Topology: {len(topology.get('compute',[]))} compute, "
          f"{len(topology.get('databases',[]))} db, "
          f"{len(topology.get('storage',[]))} storage, "
          f"{len(topology.get('network',[]))} network, "
          f"{len(topology.get('security',[]))} security nodes")
    
    return blueprint


# ── Helpers ────────────────────────────────────────────────────────

def _clean_name(name: str) -> str:
    """Normalize server/resource names for topology."""
    if not name:
        return "unnamed-resource"
    name = str(name).strip()
    name = re.sub(r'[\s_\.]+', '-', name)
    name = re.sub(r'[^a-zA-Z0-9\-]', '', name)
    return name.lower() if name else "unnamed-resource"


def _classify_network_type(item: Dict[str, Any]) -> str:
    """Classify network item into a specific type."""
    combined = f"{item.get('type', '')} {item.get('description', '')}".lower()
    if 'nat gateway' in combined:
        return 'NAT'
    elif 'virtual private network' in combined or 'vpn' in combined:
        return 'VPN'
    elif 'elastic ip' in combined or 'eip' in combined:
        return 'EIP'
    elif 'vpc' in combined:
        return 'VPC'
    elif 'direct connect' in combined:
        return 'DirectConnect'
    elif 'cdn' in combined:
        return 'CDN'
    return 'Network'


def _classify_storage_type(item: Dict[str, Any]) -> str:
    """Classify storage item into a specific type."""
    combined = f"{item.get('type', '')} {item.get('description', '')}".lower()
    if 'backup' in combined:
        return 'CBR'
    elif 'object storage' in combined or 'obs' in combined:
        return 'OBS'
    elif 'elastic volume' in combined or 'evs' in combined:
        return 'EVS'
    elif 'sfs' in combined or 'scalable file' in combined:
        return 'SFS'
    return 'Storage'
